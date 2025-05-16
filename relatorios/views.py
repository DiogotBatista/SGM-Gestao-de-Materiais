# relatorios/views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from movimentacoes.models import MovimentoItem
from .forms import RelatorioObraForm, RelatorioContratoForm, RelatorioDataForm, RelatorioMaterialForm
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import io
from materiais.models import Material


@login_required
def painel_relatorios(request):
    tipo = request.GET.get('tipo_relatorio')
    itens = None
    form = None
    page_obj = None
    consulta_realizada = False
    export_url = f"/relatorios/exportar_excel?{request.GET.urlencode()}"
    export_pdf_url = f"/relatorios/exportar_pdf?{request.GET.urlencode()}"

    if tipo == 'obra':
        form = RelatorioObraForm(request.GET)
        if form.is_valid():
            obra = form.cleaned_data['obra']
            itens = MovimentoItem.objects.filter(
                movimentacao__obra=obra,
                tipo='SAI'
            ).select_related('material', 'movimentacao', 'movimentacao__obra', 'movimentacao__obra__contrato')
            consulta_realizada = True

    elif tipo == 'contrato':
        form = RelatorioContratoForm(request.GET)
        if form.is_valid():
            contrato = form.cleaned_data['contrato']
            itens = MovimentoItem.objects.filter(
                movimentacao__obra__contrato=contrato,
                tipo='SAI'
            ).select_related('material', 'movimentacao', 'movimentacao__obra', 'movimentacao__obra__contrato')
            consulta_realizada = True

    elif tipo == 'data':
        form = RelatorioDataForm(request.GET)
        if form.is_valid():
            data_inicio = form.cleaned_data['data_inicio']
            data_fim = form.cleaned_data['data_fim']
            itens = MovimentoItem.objects.filter(
                movimentacao__data_movimentacao__date__range=(data_inicio, data_fim)
            ).select_related('material', 'movimentacao', 'movimentacao__obra', 'movimentacao__obra__contrato')
            consulta_realizada = True

    elif tipo == 'material':
        form = RelatorioMaterialForm(request.GET)
        if form.is_valid():
            material_input = form.cleaned_data['material']
            tipo_mov = form.cleaned_data.get('tipo')

            # Tenta extrair o ID do material no formato "123 - Parafuso"
            import re
            match = re.match(r"(\d+)", material_input)
            material_id = int(match.group(1)) if match else None

            material = Material.objects.filter(id=material_id).first()

            if material:
                itens = MovimentoItem.objects.filter(material=material)
                if tipo_mov:
                    itens = itens.filter(tipo=tipo_mov)
                itens = itens.select_related('movimentacao', 'movimentacao__obra', 'movimentacao__obra__contrato')
                consulta_realizada = True
            else:
                form.add_error('material', 'Material não encontrado.')

    if itens is not None:
        itens = itens.order_by('-movimentacao__data_movimentacao')
        paginator = Paginator(itens, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']

    context = {
        'tipo': tipo,
        'form': form,
        'page_obj': page_obj,
        'consulta_realizada': consulta_realizada,
        'query_params': query_params.urlencode(),
        'total_resultados': itens.count() if itens is not None else 0,
        'mostrar_tipo': tipo == 'data' or tipo == 'material',
        'mostrar_documento': tipo == 'material' and request.GET.get('tipo') == 'ENT',
        'mostrar_obra_contrato': tipo == 'material' and request.GET.get('tipo') == 'SAI',
        'mostrar_todos_os_tipos': tipo == 'material' and request.GET.get('tipo') == '',
        'export_url': export_url,
        'export_pdf_url': export_pdf_url,
    }
    return render(request, 'relatorios/painel.html', context)

@login_required
def exportar_excel(request):
    tipo = request.GET.get('tipo_relatorio')
    itens = []
    form = None

    if tipo == 'obra':
        form = RelatorioObraForm(request.GET)
        if form.is_valid():
            obra = form.cleaned_data['obra']
            itens = MovimentoItem.objects.filter(
                movimentacao__obra=obra,
                tipo='SAI'
            )

    elif tipo == 'contrato':
        form = RelatorioContratoForm(request.GET)
        if form.is_valid():
            contrato = form.cleaned_data['contrato']
            itens = MovimentoItem.objects.filter(
                movimentacao__obra__contrato=contrato,
                tipo='SAI'
            )

    elif tipo == 'data':
        form = RelatorioDataForm(request.GET)
        if form.is_valid():
            data_inicio = form.cleaned_data['data_inicio']
            data_fim = form.cleaned_data['data_fim']
            itens = MovimentoItem.objects.filter(
                movimentacao__data_movimentacao__date__range=(data_inicio, data_fim)
            )

    elif tipo == 'material':
        form = RelatorioMaterialForm(request.GET)
        if form.is_valid():
            material = form.cleaned_data['material']
            tipo_mov = form.cleaned_data.get('tipo')
            itens = MovimentoItem.objects.filter(material=material)
            if tipo_mov:
                itens = itens.filter(tipo=tipo_mov)

    itens = itens.order_by('-movimentacao__data_movimentacao')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Relatório'
    cabecalhos = ['Material', 'Quantidade', 'Tipo', 'Doc.', 'Obra', 'Contrato', 'Data da Movimentação', 'Código']
    ws.append(cabecalhos)

    for item in itens:
        ws.append([
            item.material.nome,
            item.quantidade,
            item.get_tipo_display(),
            item.movimentacao.documento or '-',
            str(item.movimentacao.obra) if item.movimentacao.obra else '-',
            str(item.movimentacao.obra.contrato) if item.movimentacao.obra and item.movimentacao.obra.contrato else '-',
            item.movimentacao.data_movimentacao.strftime('%d/%m/%Y %H:%M'),
            item.movimentacao.codigo
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=relatorio_materiais.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_pdf(request):
    tipo = request.GET.get('tipo_relatorio')
    itens = []
    form = None

    if tipo == 'obra':
        form = RelatorioObraForm(request.GET)
        if form.is_valid():
            obra = form.cleaned_data['obra']
            itens = MovimentoItem.objects.filter(movimentacao__obra=obra, tipo='SAI')
    elif tipo == 'contrato':
        form = RelatorioContratoForm(request.GET)
        if form.is_valid():
            contrato = form.cleaned_data['contrato']
            itens = MovimentoItem.objects.filter(movimentacao__obra__contrato=contrato, tipo='SAI')
    elif tipo == 'data':
        form = RelatorioDataForm(request.GET)
        if form.is_valid():
            data_inicio = form.cleaned_data['data_inicio']
            data_fim = form.cleaned_data['data_fim']
            itens = MovimentoItem.objects.filter(movimentacao__data_movimentacao__date__range=(data_inicio, data_fim))
    elif tipo == 'material':
        form = RelatorioMaterialForm(request.GET)
        if form.is_valid():
            material = form.cleaned_data['material']
            tipo_mov = form.cleaned_data.get('tipo')
            itens = MovimentoItem.objects.filter(material=material)
            if tipo_mov:
                itens = itens.filter(tipo=tipo_mov)

    itens = itens.order_by('-movimentacao__data_movimentacao')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elementos = []
    style = getSampleStyleSheet()['Normal']

    data = [['Material', 'Quantidade', 'Tipo', 'Doc.', 'Obra', 'Contrato', 'Data', 'Código']]
    for item in itens:
        data.append([
            item.material.nome,
            item.quantidade,
            item.get_tipo_display(),
            item.movimentacao.documento or '-',
            str(item.movimentacao.obra or '-'),
            str(item.movimentacao.obra.contrato) if item.movimentacao.obra and item.movimentacao.obra.contrato else '-',
            item.movimentacao.data_movimentacao.strftime('%d/%m/%Y %H:%M'),
            item.movimentacao.codigo
        ])

    tabela = Table(data, repeatRows=1)
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elementos.append(tabela)
    doc.build(elementos)

    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf', headers={
        'Content-Disposition': 'attachment; filename="relatorio_materiais.pdf"'
    })
